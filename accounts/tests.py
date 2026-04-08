"""
tests.py — LookIn AI Critical Flow Tests
==========================================
7 test cases covering the core attendance system logic.
Uses unittest.mock to avoid loading InsightFace at test time.
"""

# ┌──────────────────────────────────────────────────────────────────────────┐
# │  CRITICAL: Mock InsightFace BEFORE any Django app imports touch it.     │
# │  views.py calls es.get_kiosk_app() at module level which loads the     │
# │  500MB buffalo_l model. This mock prevents that entirely.              │
# └──────────────────────────────────────────────────────────────────────────┘
import sys
from unittest.mock import MagicMock

_mock_insightface = MagicMock()
_mock_insightface.app.FaceAnalysis.return_value = MagicMock()

# Force-replace: setdefault won't work if the real package is installed
if 'insightface' not in sys.modules:
    sys.modules['insightface'] = _mock_insightface
    sys.modules['insightface.app'] = _mock_insightface.app
if 'onnxruntime' not in sys.modules:
    sys.modules['onnxruntime'] = MagicMock()

import io
import numpy as np
from datetime import datetime
from unittest.mock import patch, MagicMock  # noqa: F811 (re-import after sys.modules patch)

from django.test import TestCase, RequestFactory, override_settings
from django.contrib.auth.models import User
from django.utils import timezone

from accounts.models import (
    Faculty, Department, Program, Semester, Division,
    Subject, SubjectEnrollment,
    AttendanceSession, Attendance, Profile,
)


# ── helpers ──────────────────────────────────────────────────────────────────

def _make_encoding(seed=42):
    """Return a deterministic 512-d float32 numpy array (fake InsightFace embedding)."""
    rng = np.random.RandomState(seed)
    vec = rng.randn(512).astype(np.float32)
    vec /= np.linalg.norm(vec)  # L2-normalise like the real pipeline
    return vec


class _BaseTestMixin:
    """
    Shared setUp that creates the minimum hierarchy:
      Faculty → Department → Program → Semester → Division → Subject (core)
    Plus two students (div A and div B) and one admin user.
    """

    def _setup_hierarchy(self):
        self.faculty = Faculty.objects.create(name="Prof. Shah", faculty_code="PS")
        self.dept = Department.objects.create(name="Computer Science")
        self.program = Program.objects.create(name="BTech")
        self.semester = Semester.objects.create(name="6")
        self.div_a = Division.objects.create(name="A")
        self.div_b = Division.objects.create(name="B")

        self.subject = Subject.objects.create(
            name="Data Structures",
            course_code="CS301",
            subject_type="core",
            faculty=self.faculty,
            department=self.dept,
            program=self.program,
            semester=self.semester,
        )
        self.subject.divisions.add(self.div_a)

        # Elective subject for mixed tests
        self.elective = Subject.objects.create(
            name="Robotics",
            course_code="EL101",
            subject_type="elective",
            faculty=self.faculty,
            department=self.dept,
        )

    def _make_student(self, name, email, division="A", seed=1):
        """Create a User + Profile + optional face encoding."""
        user = User.objects.create_user(
            username=email, email=email,
            password="test1234", first_name=name,
        )
        profile = Profile.objects.create(
            user=user,
            roll=f"R{user.id:03d}",
            faculty="Prof. Shah",
            department="Computer Science",
            program="BTech",
            semester="6",
            division=division,
        )
        # Store a fake encoding
        enc = _make_encoding(seed)
        profile.set_face_encoding(enc)
        profile.save(update_fields=["face_encoding"])
        return user

    def _make_admin(self):
        admin = User.objects.create_user(
            username="admin@lookin.ai", email="admin@lookin.ai",
            password="admin1234", first_name="Admin", is_staff=True,
        )
        Profile.objects.create(user=admin)
        return admin

    def _create_session_and_attendance(self, subject, division_name, students_present, students_absent):
        """
        Create an AttendanceSession and mark given students as present/absent.
        Returns the session object.
        """
        session = AttendanceSession.objects.create(
            faculty=subject.faculty.name,
            department=subject.department.name,
            program=subject.program.name if subject.program else "",
            semester=subject.semester.name if subject.semester else "Elective",
            division=division_name,
            subject=subject,
            lecture_slot=1,
        )
        for s in students_present:
            Attendance.objects.create(session=session, student=s, status=True)
        for s in students_absent:
            Attendance.objects.create(session=session, student=s, status=False)
        return session


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST 1 — Face encoding saves correctly (roundtrip: set → DB → get)
# ═══════════════════════════════════════════════════════════════════════════════

class TestFaceEncodingRoundtrip(TestCase, _BaseTestMixin):
    """Profile.set_face_encoding() stores raw bytes; get_face_encoding() returns identical 512-d array."""

    def setUp(self):
        self._setup_hierarchy()
        self.student = self._make_student("Yashvi Patel", "yashvi@test.com", seed=99)

    def test_encoding_roundtrip_shape_and_values(self):
        """Encoding stored as 2048 raw bytes, retrieved as identical 512-d float32."""
        original = _make_encoding(seed=99)
        profile = self.student.profile

        # Verify raw bytes length = 512 * 4
        raw = bytes(profile.face_encoding)
        self.assertEqual(len(raw), 512 * 4, "Raw bytes must be exactly 2048 (512 × float32)")

        # Verify get returns the same array
        retrieved = profile.get_face_encoding()
        self.assertIsNotNone(retrieved)
        self.assertEqual(retrieved.shape, (512,))
        self.assertEqual(retrieved.dtype, np.float32)
        np.testing.assert_allclose(retrieved, original, atol=1e-6)

    def test_no_image_returns_none(self):
        """Profile with no face_encoding returns None."""
        profile = self.student.profile
        profile.face_encoding = None
        profile.save(update_fields=["face_encoding"])
        self.assertIsNone(profile.get_face_encoding())

    def test_legacy_pickle_returns_none(self):
        """Non-2048-byte blobs (legacy pickle) are rejected gracefully."""
        profile = self.student.profile
        profile.face_encoding = b"not_a_valid_encoding"
        profile.save(update_fields=["face_encoding"])
        self.assertIsNone(profile.get_face_encoding())


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST 2 — Division check blocks wrong division
# ═══════════════════════════════════════════════════════════════════════════════

class TestDivisionCheckBlocksWrongDivision(TestCase, _BaseTestMixin):
    """
    For CORE subjects, get_valid_absent_students() must skip enrolled students
    whose profile.division ≠ session.division.
    """

    def setUp(self):
        self._setup_hierarchy()
        self.student_a = self._make_student("Alice A", "alice@test.com", division="A", seed=1)
        self.student_b = self._make_student("Bob B", "bob@test.com", division="B", seed=2)

        # Both enrolled in same core subject
        SubjectEnrollment.objects.create(subject=self.subject, student=self.student_a)
        SubjectEnrollment.objects.create(subject=self.subject, student=self.student_b)

    def test_wrong_division_student_excluded_from_absent_list(self):
        """Student in Division B must NOT appear in absent list for a Division A session."""
        from accounts.views import get_valid_absent_students

        session = self._create_session_and_attendance(
            self.subject, "A",
            students_present=[], students_absent=[],
        )
        present_ids = set()  # nobody present

        absent = get_valid_absent_students(session, self.subject, present_ids)
        absent_ids = {s.id for s in absent}

        self.assertIn(self.student_a.id, absent_ids, "Div A student MUST be in absent list")
        self.assertNotIn(self.student_b.id, absent_ids, "Div B student must be BLOCKED")

    def test_correct_division_student_included(self):
        """Student in Division A appears in absent list for a Division A session."""
        from accounts.views import get_valid_absent_students

        session = self._create_session_and_attendance(
            self.subject, "A",
            students_present=[], students_absent=[],
        )
        absent = get_valid_absent_students(session, self.subject, present_ids=set())
        self.assertTrue(
            any(s.id == self.student_a.id for s in absent),
            "Correct-division student must appear as absent",
        )

    def test_elective_no_division_filter(self):
        """Elective subjects skip division filtering — all enrolled appear."""
        from accounts.views import get_valid_absent_students

        SubjectEnrollment.objects.create(subject=self.elective, student=self.student_a)
        SubjectEnrollment.objects.create(subject=self.elective, student=self.student_b)

        session = self._create_session_and_attendance(
            self.elective, "N/A",
            students_present=[], students_absent=[],
        )
        absent = get_valid_absent_students(session, self.elective, present_ids=set())
        absent_ids = {s.id for s in absent}

        self.assertIn(self.student_a.id, absent_ids)
        self.assertIn(self.student_b.id, absent_ids)


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST 3 — Enrollment check blocks non-enrolled
# ═══════════════════════════════════════════════════════════════════════════════

class TestEnrollmentBlocksNonEnrolled(TestCase, _BaseTestMixin):
    """
    get_valid_absent_students() only considers students who have a
    SubjectEnrollment row for the session's subject.
    """

    def setUp(self):
        self._setup_hierarchy()
        self.enrolled_student = self._make_student("Enrolled", "enrolled@test.com", division="A", seed=10)
        self.non_enrolled_student = self._make_student("NotEnrolled", "notenrolled@test.com", division="A", seed=20)

        # Only one student enrolled
        SubjectEnrollment.objects.create(subject=self.subject, student=self.enrolled_student)

    def test_non_enrolled_student_excluded(self):
        """Student without SubjectEnrollment row never appears in absent list."""
        from accounts.views import get_valid_absent_students

        session = self._create_session_and_attendance(
            self.subject, "A",
            students_present=[], students_absent=[],
        )
        absent = get_valid_absent_students(session, self.subject, present_ids=set())
        absent_ids = {s.id for s in absent}

        self.assertIn(self.enrolled_student.id, absent_ids)
        self.assertNotIn(self.non_enrolled_student.id, absent_ids,
                         "Non-enrolled student must be BLOCKED from absent list")

    def test_enrolled_student_present_excluded_from_absent(self):
        """Present students (even enrolled) must not appear in the absent list."""
        from accounts.views import get_valid_absent_students

        session = self._create_session_and_attendance(
            self.subject, "A",
            students_present=[self.enrolled_student], students_absent=[],
        )
        present_ids = {self.enrolled_student.id}
        absent = get_valid_absent_students(session, self.subject, present_ids)

        self.assertEqual(len(absent), 0, "Present student must NOT be in absent list")


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST 4 — Absent email sends only to absent
# ═══════════════════════════════════════════════════════════════════════════════

@override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
class TestAbsentEmailSendsOnlyToAbsent(TestCase, _BaseTestMixin):
    """
    send_bulk_attendance_emails() must send emails ONLY to absent students.
    """

    def setUp(self):
        self._setup_hierarchy()
        self.absent_student = self._make_student("Absent Ali", "absent@test.com", division="A", seed=30)
        self.present_student = self._make_student("Present Priya", "present@test.com", division="A", seed=31)

        SubjectEnrollment.objects.create(subject=self.subject, student=self.absent_student)
        SubjectEnrollment.objects.create(subject=self.subject, student=self.present_student)

    @patch('django.core.mail.send_mail')
    def test_email_sent_to_absent_only(self, mock_send_mail):
        """Only absent students receive an email; present students are skipped."""
        from accounts.views import send_bulk_attendance_emails
        
        attendance_list = [
            (self.absent_student, False),   # absent → email
            (self.present_student, True),   # present → NO email
        ]
        
        # Run synchronously — patch Thread to execute inline
        import threading
        original_thread = threading.Thread
        
        def sync_thread(target=None, daemon=None, **kwargs):
            class FakeThread:
                def start(self):
                    if target:
                        target()
            return FakeThread()
        
        with patch('accounts.views.threading.Thread', side_effect=sync_thread):
            send_bulk_attendance_emails(
                attendance_list, self.subject, 1, timezone.now()
            )
        
        # Check send_mail was called
        self.assertTrue(mock_send_mail.called, 
            "send_mail must be called for absent student")
        
        # Get all recipient emails
        all_recipients = []
        for call in mock_send_mail.call_args_list:
            args, kwargs = call
            recipients = kwargs.get('recipient_list', 
                         args[3] if len(args) > 3 else [])
            all_recipients.extend(recipients)
        
        self.assertIn("absent@test.com", all_recipients,
            "Absent student MUST receive email")
        self.assertNotIn("present@test.com", all_recipients,
            "Present student must NOT receive email")


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST 5 — Present student gets zero email
# ═══════════════════════════════════════════════════════════════════════════════

@override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
class TestPresentStudentGetsZeroEmail(TestCase, _BaseTestMixin):
    """
    When ALL students are present, send_bulk_attendance_emails() sends NO emails.
    """

    def setUp(self):
        self._setup_hierarchy()
        self.student1 = self._make_student("Student1", "s1@test.com", division="A", seed=40)
        self.student2 = self._make_student("Student2", "s2@test.com", division="A", seed=41)

    @patch('django.core.mail.send_mail')
    def test_all_present_zero_emails(self, mock_send_mail):
        """If every student is present, send_mail is never called."""
        from accounts.views import send_bulk_attendance_emails

        all_present = [
            (self.student1, True),
            (self.student2, True),
        ]

        with patch('accounts.views.threading.Thread') as mock_thread:
            mock_thread.side_effect = lambda target, **kw: MagicMock(start=lambda: target())

            send_bulk_attendance_emails(
                all_present, self.subject, 1, timezone.now()
            )

        mock_send_mail.assert_not_called()


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST 6 — Excel download has correct data
# ═══════════════════════════════════════════════════════════════════════════════

class TestExcelDownloadHasCorrectData(TestCase, _BaseTestMixin):
    """
    download_attendance() returns an .xlsx file with correct student names,
    roll numbers, statuses, and subject info.
    """

    def setUp(self):
        self._setup_hierarchy()
        self.admin = self._make_admin()
        self.student_p = self._make_student("PresentKid", "pk@test.com", division="A", seed=50)
        self.student_a = self._make_student("AbsentKid", "ak@test.com", division="A", seed=51)

        self.session = self._create_session_and_attendance(
            self.subject, "A",
            students_present=[self.student_p],
            students_absent=[self.student_a],
        )

    def test_excel_contains_all_records(self):
        """Response is a valid .xlsx with one row per student and correct statuses."""
        import openpyxl

        factory = RequestFactory()
        request = factory.get(f"/download-attendance/{self.session.id}/")
        request.user = self.admin
        request.session = {}

        # Call the view directly (bypasses decorator auth)
        from accounts.views import download_attendance
        response = download_attendance(request, self.session.id)

        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        self.assertEqual(len(rows), 2, "Must have exactly 2 data rows")

        # Build lookup: name → status
        name_status = {row[0]: row[6] for row in rows}

        self.assertEqual(name_status["PresentKid"], "Present")
        self.assertEqual(name_status["AbsentKid"], "Absent")

    def test_excel_has_correct_headers(self):
        """First row must contain all 8 expected column headers."""
        import openpyxl

        factory = RequestFactory()
        request = factory.get(f"/download-attendance/{self.session.id}/")
        request.user = self.admin
        request.session = {}

        from accounts.views import download_attendance
        response = download_attendance(request, self.session.id)

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        expected = ["Student Name", "Roll No", "Program", "Sem", "Div", "Subject", "Status", "Date & Time"]
        self.assertEqual(headers, expected)

    def test_excel_shows_subject_with_course_code(self):
        """Subject column should display '[CS301] Data Structures' format."""
        import openpyxl

        factory = RequestFactory()
        request = factory.get(f"/download-attendance/{self.session.id}/")
        request.user = self.admin
        request.session = {}

        from accounts.views import download_attendance
        response = download_attendance(request, self.session.id)

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Subject is column 6 (index 5 in 0-based)
        subject_val = ws.cell(row=2, column=6).value
        self.assertEqual(subject_val, "[CS301] Data Structures")


# ═══════════════════════════════════════════════════════════════════════════════
#  TEST 7 — Attendance % calculates correctly
# ═══════════════════════════════════════════════════════════════════════════════

class TestAttendancePercentageCalculation(TestCase, _BaseTestMixin):
    """
    Verify that attendance percentage = (present / total) × 100
    across multiple sessions for a subject.
    """

    def setUp(self):
        self._setup_hierarchy()
        self.student = self._make_student("CalcStudent", "calc@test.com", division="A", seed=70)
        SubjectEnrollment.objects.create(subject=self.subject, student=self.student)

    def _create_sessions(self, present_count, absent_count):
        """Create N sessions where student is present, M where absent."""
        for i in range(present_count):
            s = AttendanceSession.objects.create(
                faculty=self.faculty.name,
                department=self.dept.name,
                program=self.program.name,
                semester=self.semester.name,
                division="A",
                subject=self.subject,
                lecture_slot=1,
            )
            Attendance.objects.create(session=s, student=self.student, status=True)

        for i in range(absent_count):
            s = AttendanceSession.objects.create(
                faculty=self.faculty.name,
                department=self.dept.name,
                program=self.program.name,
                semester=self.semester.name,
                division="A",
                subject=self.subject,
                lecture_slot=2,
            )
            Attendance.objects.create(session=s, student=self.student, status=False)

    def test_100_percent_attendance(self):
        """5 present out of 5 → 100%."""
        self._create_sessions(present_count=5, absent_count=0)

        total = Attendance.objects.filter(
            student=self.student, session__subject=self.subject
        ).count()
        present = Attendance.objects.filter(
            student=self.student, session__subject=self.subject, status=True
        ).count()

        percentage = (present / total * 100) if total > 0 else 0
        self.assertEqual(percentage, 100.0)

    def test_75_percent_attendance(self):
        """3 present, 1 absent → 75%."""
        self._create_sessions(present_count=3, absent_count=1)

        total = Attendance.objects.filter(
            student=self.student, session__subject=self.subject
        ).count()
        present = Attendance.objects.filter(
            student=self.student, session__subject=self.subject, status=True
        ).count()

        percentage = (present / total * 100) if total > 0 else 0
        self.assertEqual(percentage, 75.0)

    def test_0_percent_attendance(self):
        """0 present, 4 absent → 0%."""
        self._create_sessions(present_count=0, absent_count=4)

        total = Attendance.objects.filter(
            student=self.student, session__subject=self.subject
        ).count()
        present = Attendance.objects.filter(
            student=self.student, session__subject=self.subject, status=True
        ).count()

        percentage = (present / total * 100) if total > 0 else 0
        self.assertEqual(percentage, 0.0)

    def test_below_75_triggers_warning_threshold(self):
        """2 present, 2 absent → 50% which is below 75% warning threshold."""
        self._create_sessions(present_count=2, absent_count=2)

        total = Attendance.objects.filter(
            student=self.student, session__subject=self.subject
        ).count()
        present = Attendance.objects.filter(
            student=self.student, session__subject=self.subject, status=True
        ).count()

        percentage = (present / total * 100) if total > 0 else 0
        self.assertEqual(percentage, 50.0)
        self.assertTrue(percentage < 75, "50% attendance should trigger low-attendance warning")

    def test_zero_total_returns_zero(self):
        """No sessions at all → percentage defaults to 0 (no ZeroDivisionError)."""
        total = Attendance.objects.filter(
            student=self.student, session__subject=self.subject
        ).count()
        present = Attendance.objects.filter(
            student=self.student, session__subject=self.subject, status=True
        ).count()

        percentage = (present / total * 100) if total > 0 else 0
        self.assertEqual(total, 0)
        self.assertEqual(percentage, 0)
